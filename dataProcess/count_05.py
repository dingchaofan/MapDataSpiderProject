import os
import string
import sys
import shutil
# 操作excel文件
import xlrd
import xlwt
import openpyxl

import time
import datetime

# 获取脚本名的两个方法 输出脚本名
# print(__file__)  
# print(sys.argv[0])

# 获取执行脚本的目录 输出绝对路径
# print(sys.path[0])   
# print(os.getcwd())

# split方法 输出文件名
# print(os.path.split(os.getcwd()))
# print(os.path.split(__file__)[-1])
# print(os.path.split(__file__)[-1].split('.')[0])
# 输出文件的真实路径和所在文件夹的路径
# print(os.path.realpath(__file__))
# print(os.path.dirname(os.path.realpath(__file__)))

# 获取文件夹名
dirname =  os.path.split(os.getcwd())[-1]

# 定义一个数字转两位字符串的函数 0 --> '00'
def number_to_twochars(time_num):
	twocharsstr = time_num
	twocharsstr = str(time_num)
	twocharsstr = twocharsstr.zfill(2)
	return twocharsstr

# 定义一个两位字符串转数字的函数 '00' --> 0
def twochars_to_number(twocharsstr):
	time_num_str = twocharsstr.lstrip('0')
	if time_num_str == '':
		time_num_str = '0'
	time_num = int(time_num_str)
	return time_num

# 检查文件夹中文件的冗余和丢失情况 返回lostdata, redundantdata这两个数组
def checkdata():

	# 丢失的数据
	lostdata = []
	# 冗余的数据的
	redundantdata = []
	filename = ""

	for hour in range(24):
		for minute in range(60):
			flag0 = 0
			flag1 = 0
			hourstr = number_to_twochars(hour)
			minutestr = number_to_twochars(minute)
			for second in range(30):
				secondstr = number_to_twochars(second)
				filename = dirname+'_'+hourstr+'-'+minutestr+'-'+secondstr+'.png'
				if os.path.isfile(filename):
					flag0 = flag0 + 1
					if flag0 >= 2:
						redundantdata.append(filename)
			for second in range(30,60):
				secondstr =number_to_twochars(second)
				filename = dirname+'_'+hourstr+'-'+minutestr+'-'+secondstr+'.png'
				if os.path.isfile(filename):
					flag1 = flag1 + 1
					if flag1 >= 2:
						redundantdata.append(filename)
			thisminutestr = dirname+'_'+hourstr+'-'+minutestr
			if flag0 >= 1:
				pass
			else:
				lostdata.append(thisminutestr + '-before.png')
			if flag1 >= 1:
				pass
			else:
				lostdata.append(thisminutestr + '-after.png')

	# 输出丢失数据
	# for i in range(len(lostdata)):
	# 	print(lostdata[i],' NO.',i)
	# print('the num of lostdata :',len(lostdata))
	# 输出冗余数据
	# for i in range(len(redundantdata)):
	# 	print(redundantdata[i],' NO.',i)
	# print('the num of redundantdata :',len(redundantdata))

	print('the num of lostdata :',len(lostdata))
	print('the num of redundantdata :',len(redundantdata))
	return lostdata, redundantdata


def reshapedata(lostdata, redundantdata):
	redundantdataPre = []
	PreData = ""
	NextData = ""
	filename = ""
	for i in range(len(redundantdata)):

		timestr = redundantdata[i].split('_')[-1].split('.')[0]
		hourstr,minutestr,secondstr= timestr.split('-')


		hour = twochars_to_number(hourstr)
		minute = twochars_to_number(minutestr)
		second = twochars_to_number(secondstr)

		# 获取同是冗余数据的那个数据 填补冗余数据两侧的缺失数据
		if second < 30:
			for isecond in range(second-20):
					isecondstr = number_to_twochars(isecond)
					filename = dirname+'_'+hourstr+'-'+minutestr+'-'+isecondstr+'.png'
					if os.path.isfile(filename):
						redundantdataPre.append(filename)
						break

			if (minute - 1) >= 0:
				PreData =  dirname+'_'+hourstr+'-'+number_to_twochars(minute - 1) + '-after.png'
			elif (hour - 1) >= 0:
				PreData =  dirname+'_'+number_to_twochars(hour - 1)+'-59' + '-after.png'
			if (PreData in lostdata):
				if os.path.isfile(filename):
					# shutil.copy(filename,PreData.replace("after","59"))
					os.rename(filename,PreData.replace("after","59"))

			NextData = dirname+'_'+hourstr+'-'+minutestr + '-after.png'
			if (NextData in lostdata):
				shutil.copy(redundantdata[i],NextData.replace("after","30"))

		else:
			for isecond in range(30 ,second-20):
					isecondstr = number_to_twochars(isecond)
					filename = dirname+'_'+hourstr+'-'+minutestr+'-'+isecondstr+'.png'
					if os.path.isfile(filename):
						redundantdataPre.append(filename)
						break
			
			PreData = dirname+'_'+hourstr+'-'+minutestr + '-before.png'
			if (PreData in lostdata):
				if os.path.isfile(filename):
					os.rename(filename,PreData.replace("before","29"))

			if (minute + 1) <= 59:
				NextData =  dirname+'_'+hourstr+'-'+number_to_twochars(minute + 1) + '-before.png'
			elif (hour + 1) <= 23:
				NextData =  dirname+'_'+number_to_twochars(hour + 1)+'-00' + '-before.png'
			if (NextData in lostdata):
				if os.path.isfile(redundantdata[i]):
					shutil.copy(redundantdata[i],NextData.replace("before","00"))


	# for i in range(len(redundantdataPre)):
	# 	print(redundantdataPre[i],' NO.',i)
	# print('the num of redundantdataPre :',len(redundantdataPre))

def humanOps(lostdata, redundantdata):
	listYes = ['y','yse','Y']
	lostName = ""

	directWrite_instruction = input('do you want to directWrite the file details? key y/n:')
	if (directWrite_instruction in listYes):
		directWrite()
		return 0

	showdetails_instruction = input('do you want to show the file details? key y/n:')
	if (showdetails_instruction in listYes):
		print('show the file details:\n')
		# 输出丢失数据
		for i in range(len(lostdata)):
			print(lostdata[i],' NO.',i)
		print('the num of lostdata :',len(lostdata))
		# 输出冗余数据
		for i in range(len(redundantdata)):
			print(redundantdata[i],' NO.',i)
		print('the num of redundantdata :',len(redundantdata))
	else:
		print('hide the file details')

	# 判断是否删除冗余数据
	if len(redundantdata) > 0:
		delete_instruction = input('do you want to delete redundantdata? key y/n:')
		if (delete_instruction in listYes):
			print('delete redundantdata')
			for i in range(len(redundantdata)):
				if os.path.isfile(redundantdata[i]):
					os.remove(redundantdata[i])
					# redundantdata.pop(i)
		else:
			print('save redundantdata for a while')

	# 是否输出缺失情况
	if len(lostdata) > 0:
		output_lostdataname = input('do you want to print lost data name? key y/n:')
		if (output_lostdataname in listYes):
			print("lostdata num is:",len(lostdata))
			for i in range(len(lostdata)):
				name = lostdata[i].split('_')[1]
				name = name.split('.')[0]
				if(len(lostName) <= 1):
					lostName = name
				else:
					lostName = lostName +'、'+ name
			print(lostName)


	write_excel_instruction = input('do you want to write excel? key y/n:')
	if (write_excel_instruction in listYes):
		write_excel(lostdata,lostName)

	input('press any key to exit')	

# 删除过小的文件
def delete_small_files(numKB = 0):
	dirPath = os.getcwd()
	fileList = os.listdir(dirPath)
	for file in fileList:
		if (os.path.splitext(file)[-1] == ".py"):
			print(file)
		elif (os.path.splitext(file)[-1] == ".png"):
			if os.path.getsize(file) < 1024*numKB:
				print(os.path.getsize(file))
				os.remove(file)
		else:
			print(file)

def write_excel(lostdata,lostName):
	# 打开xlsx
	workbook = openpyxl.load_workbook('../../数据情况mapdata.xlsx')
	# 获取工作表
	sheet = workbook[workbook.sheetnames[0]]

	datalist = dirname.split('-')
	# 从str转换到int
	for x in range(len(datalist)):
		datalist[x] = twochars_to_number(datalist[x])

	# 获取单元格中的值 openpyxl是从1开始计数的
	for data_index in range(2,sheet.max_row):
		cell_value = sheet.cell(row=data_index, column=1).value
		if (type(cell_value) == datetime.datetime):
			if (cell_value.year == datalist[0]) and (cell_value.month == datalist[1]) and (cell_value.day == datalist[2]):
				# 向excel中写入数据
				sheet.cell(row=data_index, column=3).value = len(lostdata)
				sheet.cell(row=data_index, column=4).value = lostName
				workbook.save(r'../../数据情况mapdata.xlsx')
				print("write finished")
		else:
			print("pass not datetime cell")
			pass
# 直接操作写入
def directWrite():
	lostName = ""
	# 删除冗余数据
	if len(redundantdata) > 0:
		print('delete redundantdata')
		for i in range(len(redundantdata)):
			if os.path.isfile(redundantdata[i]):
				os.remove(redundantdata[i])
				# redundantdata.pop(i)

	# 输出缺失情况
	if len(lostdata) > 0:
		print("lostdata num is:",len(lostdata))
		for i in range(len(lostdata)):
			name = lostdata[i].split('_')[1]
			name = name.split('.')[0]
			if(len(lostName) <= 1):
				lostName = name
			else:
				lostName = lostName +'、'+ name
		print(lostName)

	# 写入表格
	write_excel(lostdata,lostName)
	input('press any key to exit')

delete_small_files(650)
lostdata, redundantdata = checkdata()

if len(lostdata) > 0:
	reshapedata(lostdata, redundantdata)
	lostdata, redundantdata = checkdata()
	humanOps(lostdata, redundantdata)
else:
	print("Today's data is good")
	directWrite()




# 日期 星期 缺失数 具体的缺失情况 备注 (日期和星期是自动生成的)
# 2018/7/2 星期一 6
# EXCEL的日期是以序列数的形式存储的，即保存的日期实际是这个日期到1900-1-1相差的天数。
# 当单元格格式为数值时，就会显示出这个差值，即2006-12-1与1900-1-1相差39052天











